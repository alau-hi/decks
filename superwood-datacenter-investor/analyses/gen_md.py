import json
from pycel import ExcelCompiler
from openpyxl import load_workbook
meta=json.load(open("model_rows.json"))
xl=ExcelCompiler(filename="materials-mass-and-replacement.xlsx"); wb=load_workbook("materials-mass-and-replacement.xlsx"); M=wb[meta.get("sheet","1 GW data center")]
def ev(sheet,coord):
    v=wb[sheet][coord].value
    if isinstance(v,str) and v.startswith("="): return xl.evaluate(f"'{sheet}'!{coord}")
    return v
S=meta.get("sheet","1 GW data center"); rows=[]; HORS=("Immediate","Soon","Medium term","Long term")
for r in meta["data_rows"]:
    sh={h:float(ev(S,f"{col}{r}")) for h,col in zip(HORS,"FGHI")}
    hor=max(sh,key=sh.get) if sum(sh.values())>0 else "Not addressed"
    rows.append(dict(name=M.cell(row=r,column=1).value, lo=ev(S,f"B{r}"), hi=ev(S,f"C{r}"), clo=ev(S,f"D{r}"), chi=ev(S,f"E{r}"), hor=hor, shares=sh,
           share=sum(sh.values()), rlo=ev(S,f"J{r}"), rhi=ev(S,f"K{r}"), slo=ev(S,f"L{r}"), shi=ev(S,f"M{r}"), note=M.cell(row=r,column=14).value))
def k(v): return f"{v:,.0f}"
def kr(lo,hi): return f"{k(lo)}–{k(hi)} tons" if (lo or hi) else "—"
def pct(lo,hi):
    a,b=sorted((lo*100,hi*100)); return f"{a:.1f}%" if abs(a-b)<0.05 else f"{a:.1f}–{b:.1f}%"
def pct0(lo,hi):
    a,b=sorted((round(lo*100),round(hi*100))); return f"about {a}%" if a==b else f"{a}–{b}%"
tot_lo,tot_hi=ev(S,f"B{meta['rt']}"),ev(S,f"C{meta['rt']}"); ex_lo,ex_hi=ev(S,f"B{meta['rx']}"),ev(S,f"C{meta['rx']}")
print("TOTAL",round(tot_lo),round(tot_hi),"EX",round(ex_lo),round(ex_hi))
for d in rows: print(f"{d['name'][:52]:52} {d['lo']:>12,.0f} {d['hi']:>12,.0f}  {d['hor']:13} {d['share']:.0%} repl {d['rlo']:>9,.0f} {d['rhi']:>9,.0f} sw {d['slo']:>8,.0f} {d['shi']:>8,.0f}")
H={}
for hor in HORS:
    rr,rs,ry=meta["hor"][hor]
    H[hor]=dict(rlo=ev(S,f"B{rr}"),rhi=ev(S,f"C{rr}"),slo=ev(S,f"B{rs}"),shi=ev(S,f"C{rs}"),ylo=ev(S,f"B{ry}"),yhi=ev(S,f"C{ry}"))
    if hor=="Immediate": H[hor]["sflo"]=ev(S,f"B{rs+1}"); H[hor]["sfhi"]=ev(S,f"C{rs+1}")
    print(hor,{a:round(b,1) for a,b in H[hor].items()})
cum=dict(rlo=ev(S,f"B{meta['rc']}"),rhi=ev(S,f"C{meta['rc']}"),slo=ev(S,f"B{meta['rsw']}"),shi=ev(S,f"C{meta['rsw']}"),ylo=ev(S,f"B{meta['rsy']}"),yhi=ev(S,f"C{meta['rsy']}"),
         shlo=ev(S,f"B{meta['rsh']}"),shhi=ev(S,f"C{meta['rsh']}"),xlo=ev(S,f"B{meta['rshx']}"),xhi=ev(S,f"C{meta['rshx']}"),nlo=ev(S,f"B{meta['rnr']}"),nhi=ev(S,f"C{meta['rnr']}"),nslo=ev(S,f"B{meta['rnrs']}"),nshi=ev(S,f"C{meta['rnrs']}"))
print("CUM",{a:round(b,3) for a,b in cum.items()})
carb=[[ev("Carbon",f"{c}{i}") for c in "BCDEFGHIJK"] for i in (2,3,4,5)]
CB={k:(ev(S,f"B{meta[k]}"),ev(S,f"C{meta[k]}")) for k in ("rcb","rcs","rcc","rcss","rccs","rcas","rceaf")}
comp_carb=[dict(name=M.cell(row=r,column=1).value,kind=M.cell(row=r,column=28).value,fac=ev(S,f"X{r}"),lo=ev(S,f"Y{r}"),hi=ev(S,f"Z{r}")) for r in meta["data_rows"] if M.cell(row=r,column=2).value is not None]
SSH=wb["Steel share"]
ss=[dict(name=SSH.cell(row=r,column=1).value,lo=ev("Steel share",f"B{r}"),hi=ev("Steel share",f"C{r}"),flo=SSH.cell(row=r,column=4).value,fhi=SSH.cell(row=r,column=5).value,slo=ev("Steel share",f"F{r}"),shi=ev("Steel share",f"G{r}"),basis=SSH.cell(row=r,column=8).value) for r in meta["ss_rows"]]
SST={k:(ev("Steel share",f"B{meta[k]}"),ev("Steel share",f"C{meta[k]}")) for k in ("ss_tot","ss_steel","ss_share","ss_share_rebar","ss_se","ss_mp","ss_tu","ss_pre")}
print("CARBON",[[ (round(x) if isinstance(x,(int,float)) else x) for x in c] for c in carb])

import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
from matplotlib.patches import Patch
INK="#1f150c"; CREAM="#f4ecdf"; DIM="#cdbfa9"; MUTED="#9d8d76"; GOLD="#e2b877"; WOOD="#b87d44"; BRIGHT="#cda165"; GREEN="#8fb356"; TEAL="#5ea9a2"; NR="#3a2b1a"
col={"Immediate":GOLD,"Soon":GREEN,"Medium term":WOOD,"Long term":TEAL,"Not addressed":NR}
plt.rcParams.update({"font.family":"DejaVu Sans","text.color":CREAM,"axes.labelcolor":DIM,"xtick.color":DIM,"ytick.color":DIM,"axes.edgecolor":NR})
fig,(a1,a2)=plt.subplots(1,2,figsize=(16,7.4),gridspec_kw={"width_ratios":[1,1.6]},facecolor=INK)
for a in (a1,a2): a.set_facecolor(INK); [a.spines[s].set_visible(False) for s in ("top","right")]
# left: build-up, high case; concrete split into slab (not addressed) and foundations (long term)
labels=["Concrete","Rebar","Structural steel","Skins, screens, fences","Platforms, racking, tray, ducting, finishes","Electrical","Mechanical","IT"]
groups={"Concrete":["Concrete","Precast"],"Rebar":["Rebar"],"Structural steel":["Structural steel"],"Skins, screens, fences":["Exterior skins","Louvers","Security"],"Platforms, racking, tray, ducting, finishes":["Platforms","Acoustic","Racking","Other","Ducting","Interior"],"Electrical":["Electrical"],"Mechanical":["Mechanical"],"IT":["IT"]}
base=0
for i,g in enumerate(labels):
    grp=[d for d in rows if any(d["name"].startswith(p) for p in groups[g])]
    v=sum(d["hi"] for d in grp)/1e6; rep=sum(d["rhi"] for d in grp)/1e6
    a1.bar(i,v,bottom=base,color=NR,width=0.7)
    if rep>0:
        hor=[d["hor"] for d in grp if d["rhi"]>0]; hcol=col[max(set(hor),key=hor.count)]
        a1.bar(i,rep,bottom=base,color=hcol,width=0.7)
    a1.text(i,base+v+0.03,f"{v*1e6:,.0f} tons" if v<0.5 else f"{v:.2f} million tons",ha="center",fontsize=8.5,color=CREAM); base+=v
a1.bar(len(labels),base,color=MUTED,width=0.7); a1.text(len(labels),base+0.03,f"{base:.2f} Mt",ha="center",fontsize=9,color=CREAM,fontweight="bold")
a1.set_xticks(range(len(labels)+1)); a1.set_xticklabels(labels+["Total"],rotation=35,ha="right",fontsize=8.5); a1.set_ylabel("million tons per 1 GW data center (high case)")
a1.set_title("Mass build-up — building, then contents (colored = replaceable)",loc="left",fontsize=12,color=CREAM)
# right: everything except slab concrete
ex=[d for d in rows if not d["name"].startswith("Precast") and d["hi"]>0]
names=[d["name"].split(" (")[0].replace("Structural steel — ","Steel — ").replace("Concrete: foundations, footings, piers, equipment pads","Concrete — foundations, footings, pads").replace("Concrete: slab on grade, paving, yard","Concrete — slab on grade, paving") for d in ex]
y=list(range(len(ex)))
a2.barh(y,[d["rhi"]/1000 for d in ex],color=[col[d["hor"]] for d in ex],height=0.65)
a2.barh(y,[(d["hi"]-d["rhi"])/1000 for d in ex],left=[d["rhi"]/1000 for d in ex],color=NR,height=0.65)
xmax=max(d["hi"] for d in ex)/1000
for i,d in enumerate(ex): a2.text(d["hi"]/1000+xmax*0.01,i,f"{d['hi']:,.0f} tons"+(f" · {d['rhi']/d['hi']:.0%} {d['hor'].lower()}" if d["rhi"] else ""),va="center",fontsize=8.5,color=DIM)
a2.set_yticks(y); a2.set_yticklabels(names,fontsize=8.5); a2.invert_yaxis(); a2.set_xlabel("tons per 1 GW data center (high case) — log scale"); a2.set_xscale("log"); a2.set_xlim(0.5,xmax*6)
a2.set_title("Every component — what SUPERWOOD replaces, and when (log scale)",loc="left",fontsize=12,color=CREAM)
fig.legend(handles=[Patch(color=GOLD,label="Immediate — skins, screens, fences, interiors"),Patch(color=GREEN,label="Soon — platforms, barriers, racking, doors"),Patch(color=WOOD,label="Medium term — structural steel, roofs, ducting, enclosures"),Patch(color=TEAL,label="Long term — slab, paving, foundations (technical potential)"),Patch(color=NR,label="Not replaced")],loc="lower center",ncol=5,fontsize=8.5,facecolor=INK,edgecolor=NR,bbox_to_anchor=(0.5,0.035))
fig.text(0.01,0.01,"Estimates for a 1 GW IT-load data center; only concrete and structural-steel intensities are published (secondary, conf M). Long-term concrete shares are stated technical potential (Alex, 2026-09-01), not an engineered plan. Source model: materials-mass-and-replacement.xlsx",fontsize=8,color=MUTED)
plt.tight_layout(rect=(0,0.09,1,1)); fig.savefig("mass-buildup-and-replacement.png",dpi=160,facecolor=INK); print("png ok")

def row_md(d): return f"| {d['name']} | {kr(d['lo'],d['hi'])} | {kr(d['clo'],d['chi'])} |"
def rep_md(d):
    cells=" | ".join(f"{d['shares'][h]:.0%}" if d['shares'][h] else "—" for h in HORS)
    return f"| {d['name']} | {cells} | {kr(d['rlo'],d['rhi'])} | {kr(d['slo'],d['shi'])} | {d['note']} |"
bld=rows[:9]; cont=rows[9:]; imm,soon,med,lng=(H[h] for h in HORS)
NL="\n"
def hrow(label,h,plant):
    return f"| {label} | {kr(h['rlo'],h['rhi'])} | {kr(h['slo'],h['shi'])}{' ('+f'{h['sflo']/1e6:.1f}–{h['sfhi']/1e6:.1f}M sf)' if 'sflo' in h else ''} | {h['ylo']:.1f}–{h['yhi']:.1f} yr of {plant} |"
md=f"""# Material mass in a data center — the build-up, and how much SUPERWOOD can replace

Date: 2026-09-05 (v8: equipment rows carry an embodied-carbon estimate for their steel content only — 60% of electrical, 50% of mechanical, 40% of IT mass at the steel factor (Alex 2026-09-05); v6: per-component embodied carbon, steel and concrete carbon shares, EAF sensitivity and a steel-share-of-above-ground-mass sheet added to the workbook; v5: racking stays *Soon* — a few months of development; server and equipment enclosures, 40% of IT mass, added to the *Long term*; electronics never — per Alex 2026-09-04. v4 2026-09-01 set the long-term concrete shares). Status: **estimate**.
Every table below is generated from the live model [materials-mass-and-replacement.xlsx](materials-mass-and-replacement.xlsx)
— change an input there and regenerate rather than hand-edit. Labels: published / derived / estimated; confidence
`[conf: H|M|L]`. Treat everything as `[conf: L]` unless marked.

Reference unit: **1 GW of IT load** — a data center of roughly 5–10 hyperscale buildings on the assumptions below.
Low and high columns are whole scenarios (all-low inputs, all-high inputs), not a distribution.

![Mass build-up and replacement](mass-buildup-and-replacement.png)

## The four horizons

| Horizon | What it means | Product and plant | Gate |
|---|---|---|---|
| **Immediate** | Replacements shipping now | SuperMill One boards to 8" × 16' × 3/8" | None beyond E84 Class A where a finish rating applies |
| **Soon** | Non-structural items behind one scoped test program — racks, platforms, barriers, doors | SuperMill One, then SuperMill Two | A scoped test program per application; racks are a few months of development |
| **Medium term** | Structural steel (primary and roof), roof trusses and roofs, ducting, enclosures | SuperMill Two boards and veneers | Mass-timber qualification pathway, ICC-ES, E119, NFPA 285, FM acceptance; NFPA 90A / UL 181 for ducting |
| **Long term** | Concrete: 75% of slab-on-grade and paving, 90% of foundations, footings, piers and pads, with the rebar in each; server and equipment enclosures (40% of IT mass, estimate) — never the electronics | ChipMill-scale products | Stated technical potential (Alex, 2026-09-01 and 2026-09-04); no design or code pathway yet |

## 1. Assumptions that drive the build-up

| Item | Low | High | Label |
|---|---|---|---|
| Building area per MW | 5,000 sf | 10,000 sf | estimated |
| Buildings per GW (1M sf footprint each) | 5 | 10 | derived |
| Exterior wall area (4 × 1,000 ft × 40 ft per building) | 0.8M sf | 1.6M sf | derived |
| Perimeter walls | metal panel / IMP (switch to precast in the model) | | choice |
| Concrete intensity, all concrete | 500 m³/MW | 1,000 m³/MW | published, secondary [M] — arXiv 2509.21312 citing Hasan 2022 / Sharma 2023 |
| Share of concrete in foundations, footings, piers, pads | 40% | 50% | estimated |
| Structural steel intensity | 50 t/MW | 100 t/MW | same source [M] |
| Roof-and-secondary share of structural steel (trusses, joists, deck, girts) | 40% | 60% | estimated |
| Rebar | 60 kg/m³ | 100 kg/m³ | estimated |
| Ducting and air-distribution sheet metal | 3 t/MW | 8 t/MW | estimated |
| Electrical / mechanical / IT | 50 / 20 / 15 t/MW | 100 / 50 / 70 t/MW | estimated from unit masses; GB200 NVL72 rack 1.36 t [H] |
| SUPERWOOD board | 7.2 mm × 1.3 t/m³ = 0.87 kg/sf; SuperMill One ≈ 900 tons/yr, SuperMill Two ≈ 31,000 tons/yr | | internal TEM basis |
| Substitution, steel | 0.3 kg SUPERWOOD per kg steel | 0.6 kg/kg | estimated — to be engineering-stamped per element |
| Long-term technical potential, concrete | 75% of slab and paving; 90% of foundations, footings, piers, pads | same | asserted-internal (Alex, 2026-09-01) |
| Substitution, concrete | 0.05 kg SUPERWOOD per kg concrete | 0.15 kg/kg | estimated — lightweight insulated wood-foundation concept, no design exists |

## 2. The build-up: what goes into a 1 GW data center

Read top to bottom. Building structure and envelope first, then contents. The cumulative column is the running total.

### 2a. Building — structure and envelope

| Component | Mass per data center | Cumulative |
|---|---|---|
{NL.join(row_md(d) for d in bld)}

### 2b. Contents — fit-out and equipment

| Component | Mass per data center | Cumulative |
|---|---|---|
{NL.join(row_md(d) for d in cont)}

**Total mass, building and contents: {kr(tot_lo,tot_hi)}** (about {tot_lo/1e6:.1f}–{tot_hi/1e6:.1f} Mt).
**Excluding concrete: {kr(ex_lo,ex_hi)}** — concrete is {pct0(1-ex_lo/tot_lo,1-ex_hi/tot_hi)} of everything by mass.

Two consequences follow before any SUPERWOOD arithmetic:

- Any "share of a data center" statement must say whether concrete is in the denominator.
- Excluding concrete, steel in all forms (structural, rebar, roughly 60% of equipment mass as enclosures and cores,
  racks, tray, ducting) comes out near 70% of what is left on these assumptions — consistent in order of magnitude
  with the company's ex-concrete "~60% steel" framing, but a derivation, not a source.

## 3. How much of it SUPERWOOD can replace — component by component

Same rows as Section 2, with the share of each component SUPERWOOD can address in each horizon (a row may split across
horizons; these are the editable columns on the workbook sheet *1 GW data center*), the incumbent mass replaced, and
the SUPERWOOD mass required.

### 3a. Building — structure and envelope

| Component | Immediate | Soon | Medium term | Long term | Incumbent replaced | SUPERWOOD required | Gate / basis |
|---|---|---|---|---|---|---|---|
{NL.join(rep_md(d) for d in bld)}

### 3b. Contents — fit-out and equipment

| Component | Immediate | Soon | Medium term | Long term | Incumbent replaced | SUPERWOOD required | Gate / basis |
|---|---|---|---|---|---|---|---|
{NL.join(rep_md(d) for d in cont)}

## 4. Roll-up by horizon (per 1 GW data center)

| Horizon | Incumbent mass replaced | SUPERWOOD required | Plant-years |
|---|---|---|---|
{hrow("Immediate — skins, screens, fences, interiors, backplanes",imm,"SuperMill One")}
{hrow("Soon — platforms, railings, barriers, racking, doors",soon,"SuperMill Two")}
{hrow("Medium term — structural steel, roof trusses and roofs, ducting, enclosures",med,"SuperMill Two")}
{hrow("Long term — slab, paving, foundations and their rebar (technical potential)",lng,"SuperMill Two")}
| **Cumulative** | **{kr(cum['rlo'],cum['rhi'])}** | **{kr(cum['slo'],cum['shi'])}** | **{cum['ylo']:.1f}–{cum['yhi']:.1f} yr of SuperMill Two** |
| Not replaced by SUPERWOOD | {kr(cum['nlo'],cum['nhi'])} | | {pct0(cum['nslo'],cum['nshi'])} of total mass |

- Through the medium term SUPERWOOD addresses **{kr(imm['rlo']+soon['rlo']+med['rlo'],imm['rhi']+soon['rhi']+med['rhi'])}** of incumbent
  material — essentially all the steel above the slab, about
  {pct((imm['rlo']+soon['rlo']+med['rlo'])/tot_lo,(imm['rhi']+soon['rhi']+med['rhi'])/tot_hi)} of total data center mass.
  The long-term concrete rows are what move the total: with them, the ceiling is **{pct(cum['shlo'],cum['shhi'])} of total
  data center mass**. Those shares (75% of slab and paving, 90% of foundations) are stated technical potential, not an
  engineered plan, and the whole difference between the two figures rests on them.
- What stays: the remaining concrete, servers, gensets, transformers, switchgear, batteries, chillers, copper, loop
  water — {pct0(cum['nslo'],cum['nshi'])} of total mass.
- Plant math: one gigawatt data center's immediate skins are {imm['sflo']/1e6:.1f}–{imm['sfhi']/1e6:.1f}M sf, **{imm['ylo']:.1f}–{imm['yhi']:.1f} years
  of SuperMill One's entire output**. The medium-term structural horizon alone is {med['ylo']:.1f}–{med['yhi']:.1f} years of SuperMill
  Two per gigawatt — two or three gigawatts of data center absorb the plant for years, which is the offtake argument and
  the capacity risk in one number. The long-term concrete rows, if they ever became real, are a ChipMill-scale market on their own.

## 5. Embodied-carbon effect (pre-LCA, per GW)

Deck figures: steel 1.8 kg CO₂e/kg (global BF-BOF average), concrete 0.12 kg/kg, SUPERWOOD 0.5 kg/kg manufactured and
1.3 kg/kg biogenic carbon stored — **pre-LCA projections at scale; LCA under way with Prof. Ming Hu, University of Notre
Dame.** Biogenic storage reported separately (EN 15804 module C). Each component is valued at its own factor: concrete
rows at 0.12, steel rows (including rebar) at 1.8, interior finishes at 1.0 [estimate]; equipment rows, including the
server enclosures in the long-term horizon, carry no factor, so long-term avoided emissions count concrete and rebar only. The "vs EAF" column revalues steel at 0.7 kg/kg (recycled, high end).

| Horizon | Incumbent emissions avoided | SUPERWOOD manufacturing | Net reduction | Net vs EAF steel | Biogenic stored (separate) |
|---|---|---|---|---|---|
""" + NL.join(f"| {h} | {kr(c[0],c[1])} CO₂e | {kr(c[2],c[3])} | **{kr(c[4],c[5])}** | {kr(c[6],c[7]) if isinstance(c[6],(int,float)) else 'n/a'} | {kr(c[8],c[9])} |" for h,c in zip(("Immediate (steel)","Soon (steel)","Medium term (steel)","Long term (foundation concrete)"),carb)) + f"""

Against recycled (EAF) steel the avoided figure is far lower, so any claim must state its baseline.

### 5b. Embodied carbon by component (equipment rows: steel content only)

The slide-5 "by embodied carbon" view. Factors: concrete {comp_carb[0]['fac']} kg CO₂e/kg [M], steel 1.8 [M, global
BF-BOF average], interior finishes 1.0 [L]. Equipment (electrical, mechanical, IT) is outside a materials estimate.

| Component | Class | Factor kg CO₂e/kg | Embodied carbon, low–high |
|---|---|---|---|
""" + NL.join(f"| {d['name']} | {d['kind']} | {(f"{d['fac']:.2f} (steel content)") if d['kind']=='equipment' else d['fac']} | {(kr(d['lo'],d['hi'])+' CO₂e') if (d['lo'] or d['hi']) else '—'} |" for d in comp_carb) + f"""

| Roll-up | Low | High |
|---|---|---|
| Embodied carbon, building materials | {k(CB['rcb'][0])} tons CO₂e | {k(CB['rcb'][1])} tons CO₂e |
| Steel share | {CB['rcss'][0]:.0%} | {CB['rcss'][1]:.0%} |
| Concrete share | {CB['rccs'][0]:.0%} | {CB['rccs'][1]:.0%} |
| Steel above the slab (steel excl. rebar), share | {CB['rcas'][0]:.0%} | {CB['rcas'][1]:.0%} |
| Steel share if steel is recycled (EAF, 0.7 kg/kg) | {CB['rceaf'][0]:.0%} | {CB['rceaf'][1]:.0%} |

By embodied carbon, steel is roughly 55–60% of the building materials and concrete roughly 40–45% at the global-average
steel factor; with recycled steel the steel share falls to about a third. The steel above the slab — what SUPERWOOD
addresses through the structural horizon — carries about 37% of the building materials' embodied carbon.

### 5c. Steel share of above-ground mass, contents included (workbook sheet *Steel share*)

Basis for the slide-5 headline. Excludes slab, paving, foundations and all concrete. Steel fractions per component
are estimates [conf: L] and editable in the workbook. Printed band on the deck: **50–80%** (Alex, 2026-09-04).

| Component | Mass, low–high | Steel fraction | Steel, low–high | Basis |
|---|---|---|---|---|
""" + NL.join(f"| {d['name']} | {kr(d['lo'],d['hi'])} | {d['flo']:.0%}–{d['fhi']:.0%} | {kr(d['slo'],d['shi'])} | {d['basis']} |" for d in ss) + f"""
| **Above-ground total** | **{kr(*SST['ss_tot'])}** | | **{kr(*SST['ss_steel'])}** | |

| Share | Low | High |
|---|---|---|
| Steel share of above-ground mass (wall system per Inputs toggle) | {SST['ss_share'][0]:.0%} | {SST['ss_share'][1]:.0%} |
| — with metal-panel / IMP walls | {SST['ss_mp'][0]:.0%} | {SST['ss_mp'][1]:.0%} |
| — with tilt-up / precast concrete walls ({kr(*SST['ss_pre'])} of panel) | {SST['ss_tu'][0]:.0%} | {SST['ss_tu'][1]:.0%} |
| Including slab rebar (at-grade steel) | {SST['ss_share_rebar'][0]:.0%} | {SST['ss_share_rebar'][1]:.0%} |
| Structure and envelope only | {SST['ss_se'][0]:.0%} | {SST['ss_se'][1]:.0%} |

Equipment is 55–60% of above-ground mass, so the answer turns on how much of a genset, switchgear lineup, chiller and
server rack is steel; at 40% for all equipment the share falls to about 60–65%. The wall system is the other lever:
tilt-up or precast concrete walls add tens of thousands of tons of above-ground concrete and pull the share down by
about ten points, which is why the deck prints a 50–80% band rather than a point. Narrative: [steel-share-above-ground.md](steel-share-above-ground.md).

## 6. Sensitivities — what moves the answer most

1. **Long-term concrete.** The long-term rows are {kr(lng['rlo'],lng['rhi'])} of concrete and rebar on stated technical
   potential (75% of slab and paving, 90% of foundations) and a 0.05–0.15 substitution factor, none of it engineered.
   They are the only rows that change the total-mass share materially, and the least evidenced. Present them as
   technical potential and say so on the slide.
2. **Precast vs metal-panel walls.** Decides whether facades are a {kr(rows[6]['lo'],rows[6]['hi'])} metal replacement now or a
   concrete-panel replacement in the medium term behind NFPA 285 and E119.
3. **Structural steel intensity.** The 50–100 t/MW range is a factor of two; one real takeoff from HITT, Turner or
   Fast + Epp collapses it.
4. **Substitution factor.** 0.3–0.6 kg per kg is the whole medium-term SUPERWOOD demand number; Fast + Epp can stamp
   it per element.
5. **Ducting.** In-airstream components face NFPA 90A / UL 181 noncombustibility expectations; the July reviewers
   flagged this as the hardest sell in the set. Enclosures and separations are the safer framing.
6. **AI halls.** Higher rack density means fewer buildings per GW, less structure and skin per MW, more electrical and
   mechanical mass; the replaceable share falls.

## 7. Gaps and next steps

- One building's material takeoff (structure, envelope, foundations, fencing) from a construction partner.
- Read Hasan et al. 2022 and Sharma et al. 2023 directly; the arXiv preprint is a secondary citation.
- Fix the TEM thickness mix so plant output in tonnes is one number.
- Fast + Epp to state substitution factors for two or three elements (truss, joist, wall panel), and a first opinion
  on whether a SUPERWOOD foundation system is even a sensible engineering target.
- Confirm with Microsoft and Meta whether their standard design uses precast or metal-panel walls.

## Sources

- arXiv 2509.21312 (Sep 2025), citing Hasan et al. 2022 and Sharma et al. 2023 — structural steel 500–1,000 t and
  concrete 5,000–10,000 m³ per 10 MW [secondary, conf: M].
- NVIDIA GB200 NVL72 rack mass ~1,360 kg — NVIDIA published specification [conf: H].
- Internal: SUPERWOOD Structural and Cladding TEMs (Jan 2026); 7.2 mm average thickness basis
  (`../../investor-overview/sources/notes/superwood-unit-economics.md`).
- Internal: carbon figures per the Canva DCII deck (0.5 / −1.3 kg CO₂e per kg SUPERWOOD; steel 1.8), pre-LCA.
  Concrete 0.12 kg CO₂e/kg is a typical ready-mix cradle-to-gate value [conf: M].
- Unit masses for gensets, UPS, transformers, chillers, metal panel, fencing, ducting, rebar ratios, foundation share:
  general industry ranges, not individually sourced [conf: L]. Replace with vendor or takeoff data before external use.
"""
open("materials-mass-and-replacement.md","w").write(md); print("md written")
