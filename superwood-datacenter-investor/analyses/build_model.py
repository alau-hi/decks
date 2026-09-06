from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter as CL
from openpyxl.chart import BarChart, Reference
wb=Workbook()
INK="1F150C"; CREAM="F4ECDF"
hdr=Font(bold=True,color=CREAM); hfill=PatternFill("solid",fgColor=INK); inp=PatternFill("solid",fgColor="FFF4D6"); calc=PatternFill("solid",fgColor="F2EDE4"); tot=Font(bold=True); sect=PatternFill("solid",fgColor="E8DCC6")
def header(ws,row,cols):
    for i,c in enumerate(cols,1):
        x=ws.cell(row=row,column=i,value=c); x.font=hdr; x.fill=hfill; x.alignment=Alignment(wrap_text=True,vertical="center")
def widths(ws,w):
    for i,x in enumerate(w,1): ws.column_dimensions[CL(i)].width=x
SF_M2="0.092903"

# ---------------- Inputs ----------------
I=wb.active; I.title="Inputs"
header(I,1,["Parameter","Low","High","Unit","Label / confidence","Note"])
rows=[
("it_mw","IT load of reference data center",1000,1000,"MW","assumption","Everything downstream is linear in this"),
("sf_per_mw","Building area per MW",5000,10000,"sf/MW","estimated [L]","Hyperscale, single-story"),
("footprint_sf","Footprint per building",1e6,1e6,"sf","estimated [L]","~1,000 ft sides"),
("wall_sf_bldg","Exterior wall area per building",160000,160000,"sf","derived","4 × 1,000 ft × 40 ft eave"),
("yard_screen_sf","Yard / mechanical screens and louvers (per data center)",200000,500000,"sf","estimated [L]",""),
("fence_km","Perimeter + yard fence length",10,20,"km","estimated [L]","Data center perimeter plus transformer / generator yards"),
("fence_h_m","Fence height",2.5,2.5,"m","assumption",""),
("interior_sf","Biophilic interior paneling (admin / office wings)",100000,300000,"sf","estimated [L]",""),
("backplane_sf","Backplanes, trim, door kicks, sub-framing",50000,150000,"sf","estimated [L]",""),
("concrete_m3_mw","Concrete intensity (all concrete)",500,1000,"m³/MW","published, secondary [M]","arXiv 2509.21312 citing Hasan 2022 / Sharma 2023: 5,000–10,000 m³ per 10 MW"),
("concrete_t_m3","Concrete density",2.4,2.4,"t/m³","published [H]",""),
("walls_precast","Walls are precast? (1 = yes, 0 = metal panel)",0,0,"flag","choice","1: facade mass moves from metal skins (immediate) to precast panels (long term)"),
("precast_thk_m","Precast / tilt-up wall thickness",0.2,0.2,"m","estimated [L]","8 in"),
("steel_t_mw","Structural steel intensity",50,100,"t/MW","published, secondary [M]","Same source: 500–1,000 t per 10 MW"),
("secondary_share","Share of structural steel that is secondary (trusses, joists, girts, deck)",0.4,0.6,"share","estimated [L]","Remainder = primary frame"),
("rebar_kg_m3","Rebar per m³ of concrete",60,100,"kg/m³","estimated [L]",""),
("platform_t_mw","Platforms, walkways, mezzanines, railings, tray supports",5,15,"t/MW","estimated [L]",""),
("panel_kg_m2","Metal panel / IMP cladding mass",25,50,"kg/m²","estimated [L]",""),
("louver_kg_m2","Louver / screen system mass",10,20,"kg/m²","estimated [L]",""),
("fence_kg_m","Steel fence mass per running metre",30,50,"kg/m","estimated [L]","Palisade / chain-link with posts"),
("acoustic_t_mw","Acoustic barriers, enclosures, HVAC separations",2,5,"t/MW","estimated [L]",""),
("racking_t_mw","Racking and equipment supports",2,5,"t/MW","estimated [L]",""),
("other_t_mw","Other tray, containment, doors, misc. metals",5,10,"t/MW","estimated [L]",""),
("interior_t_mw","Interior finishes, backplanes, trim (mass)",1,3,"t/MW","estimated [L]","Admin / office wings only"),
("elec_t_mw","Electrical: gensets, transformers, switchgear, UPS, busway, copper",50,100,"t/MW","estimated [L]","Genset 30–60 t, 1 MW UPS 10–20 t, transformer 5–8 t"),
("mech_t_mw","Mechanical: chillers, fan walls, coolers, piping, water",20,50,"t/MW","estimated [L]",""),
("mech_shell_share","Share of mechanical mass that is equipment shells, casings and fan-wall frames — SUPERWOOD in the medium term",0.25,0.25,"share","estimated [L]; Alex 2026-09-06","About half of the steel in the row; compressor and pump bodies, tubes and steel piping excluded"),
("mech_blade_share","Share of mechanical mass that is fan blades — long term, needs investigation",0.02,0.02,"share","estimated [L]; Alex 2026-09-06","Aluminum or composite blades today; SUPERWOOD blades are a research question"),
("it_t_mw","IT: servers and racks",15,70,"t/MW","estimated [L]","Loaded rack ~1 t; GB200 NVL72 ~1.36 t [H]"),
("it_rack_share","Share of IT mass that is server rack cabinets (frames, panels, doors) — substitutable soon",0.18,0.18,"share","estimated [L]; Alex 2026-09-06","A loaded rack of 1–1.4 t carries a 150–250 kg steel cabinet"),
("it_enclosure_share","Share of IT mass that is server enclosures (sheet-steel chassis) — long term",0.30,0.30,"share","estimated [M]; analyses/it-rack-material-stack.md, applied 2026-09-06","Server boxes eventually; the electronics never"),
("foundation_share","Share of all concrete that is foundations, footings, piers and equipment pads",0.4,0.5,"share","estimated [L]","Remainder is slab on grade, paving and yard"),
("slab_lt","Long-term technical potential: share of slab-on-grade and paving concrete replaceable",0.75,0.75,"share","asserted-internal (Alex, 2026-09-01) [L]","Technical potential, not a plan; no design or code pathway yet"),
("fdn_lt","Long-term technical potential: share of foundations, footings, piers and pads replaceable",0.9,0.9,"share","asserted-internal (Alex, 2026-09-01) [L]","Technical potential, not a plan"),
("conc_sub_factor","kg SUPERWOOD per kg of concrete replaced",0.05,0.15,"kg/kg","estimated [L]","Lightweight insulated wood-foundation concept; no design exists yet"),
("duct_t_mw","Ducting, plenums and air-distribution sheet metal",3,8,"t/MW","estimated [L]","Separate from the mechanical-equipment row"),
("co2_conc","Concrete emissions",0.12,0.12,"kg CO₂e/kg","industry range [M]","Ready-mix, cradle to gate; typical 0.10–0.15"),
("co2_plastic","Polymers in paints, foam cores and plastic trim (average)",3.0,3.0,"kg CO₂e/kg","industry range [M]; Alex 2026-09-05","Cradle to gate; polyolefins and PVC about 2–2.5, polyurethane foam 3–4, coatings higher"),
("board_mm","SUPERWOOD average board thickness",7.2,7.2,"mm","internal TEM basis [H]",""),
("sw_kg_m3","SUPERWOOD density",1300,1300,"kg/m³","internal TEM [H]",""),
("sm1_sf","SuperMill One output",1e6,1e6,"sf/yr","internal [H]",""),
("sm2_sf","SuperMill Two output",36e6,36e6,"sf/yr","internal [H]",""),
("sub_factor","Structural substitution: kg SUPERWOOD per kg steel replaced",0.3,0.6,"kg/kg","estimated [L]","Engineering-stamped per element (Fast + Epp) before external use"),
("hybrid_kg_m2","SUPERWOOD in a hybrid wall panel replacing precast",20,40,"kg/m²","estimated [L]","Panel design unsettled"),
("share_other","Addressable share of other tray / containment / doors (medium term)",0.5,0.5,"share","estimated [L]",""),
("rebar_frontier","Count rebar as addressable in the long term? (0/1)",0,0,"flag","choice","Rebar is a frontier market with no pathway; off by default"),
("concrete_basis","Concrete basis: 0 = published per-MW intensity, 1 = footprint bottom-up (Foundations sheet)",1,1,"switch","choice","Deck default since 2026-09-06 (Alex): footprint basis"),
("soil_case","Soil case for the footprint basis: 1 good (spread footings), 2 moderate (lean concrete, larger footings), 3 poor (piles, pile caps, structural slab)",2,2,"switch","choice","Deck default since 2026-09-06 (Alex): moderate soils; the web deck has a soils slider"),
("slab_thk_in","Hall slab on grade thickness",6,8,"in","estimated [L]",""),
("thick_share","Share of hall floor thickened (electrical rooms, dense rack rows)",0.15,0.25,"share","estimated [L]",""),
("thick_thk_in","Thickened slab thickness",10,12,"in","estimated [L]",""),
("bay_ft","Column bay",40,40,"ft","estimated [L]","Square bays"),
("footing_ft","Spread footing plan dimension",8,10,"ft","estimated [L]","Square footings"),
("footing_d_ft","Spread footing depth",2,2.5,"ft","estimated [L]",""),
("gb_w_ft","Perimeter grade beam width",2,2.5,"ft","estimated [L]",""),
("gb_d_ft","Perimeter grade beam depth",3,3.5,"ft","estimated [L]",""),
("genset_mw","Generator unit size",3,2.5,"MW","estimated [L]","Smaller units in the high case mean more pads"),
("genset_redund","Generator redundancy factor",1.1,1.25,"factor","estimated [L]","N+1 to N+25%"),
("genset_pad_cf","Generator pad volume",480,787.5,"cf","estimated [L]","40x12x1 ft to 45x14x1.25 ft"),
("xfmr_per_gen","Transformer / switchgear pads per generator",0.5,0.6,"count","estimated [L]",""),
("xfmr_pad_cf","Transformer / switchgear pad volume",300,562.5,"cf","estimated [L]","20x15x1 ft to 25x18x1.25 ft"),
("mech_pad_share","Mechanical equipment pads as a share of floor area",0.05,0.08,"share","estimated [L]","Chillers, coolers, fan walls on grade or roof"),
("mech_pad_thk_in","Mechanical pad thickness",8,10,"in","estimated [L]",""),
("site_paving_ratio","Site paving, roads, yards, parking as a multiple of building footprint",0.8,1.5,"ratio","estimated [L]",""),
("site_paving_thk_in","Site paving thickness",6,8,"in","estimated [L]",""),
("lean_share","Case 2: share of footprint over-excavated and replaced with lean concrete / CLSM",0.3,0.6,"share","estimated [L]",""),
("lean_thk_in","Case 2: lean concrete thickness",6,12,"in","estimated [L]",""),
("found_uplift","Case 2: enlargement of footings, grade beams and pads",0.3,0.3,"factor","estimated [L]",""),
("piles_per_col","Case 3: piles or drilled piers per column",3,4,"count","estimated [L]",""),
("pile_dia_ft","Case 3: pile diameter",1.5,2,"ft","estimated [L]",""),
("pile_len_ft","Case 3: pile length",40,60,"ft","estimated [L]",""),
("cap_cf","Case 3: pile cap volume per column (replaces the spread footing)",300,504,"cf","estimated [L]","10x10x3 ft to 12x12x3.5 ft"),
("gb_pier_spacing_ft","Case 3: pier spacing under grade beams",20,15,"ft","estimated [L]",""),
("struct_slab_add_in","Case 3: slab thickening to a structural slab",2,4,"in","estimated [L]",""),
("yard_mat_acres","Substation / transformer yard mats per GW",10,20,"acres","estimated [L]",""),
("yard_mat_thk_in","Yard mat thickness",18,24,"in","estimated [L]",""),
("co2_steel","Steel emissions, global average (BF-BOF)",1.8,1.8,"kg CO₂e/kg","published-class [M]",""),
("co2_eaf","Steel emissions, recycled (EAF)",0.4,0.7,"kg CO₂e/kg","industry range [M]",""),
("co2_sw","SUPERWOOD manufacturing emissions",0.5,0.5,"kg CO₂e/kg","internal, pre-LCA [L]","Canva DCII deck; LCA under way, Prof. Ming Hu, Notre Dame"),
("co2_bio","SUPERWOOD biogenic carbon stored",1.3,1.3,"kg CO₂e/kg","internal, pre-LCA [L]","Report separately; module C release under EN 15804"),
("co2_mixed","Interior finishes emissions (gypsum, wood, steel studs — mixed)",1.0,1.0,"kg CO₂e/kg","estimated [L]","Used only for the interior finishes row"),
]
R={}
for i,(k,name,lo,hi,unit,lab,note) in enumerate(rows,start=2):
    I.cell(row=i,column=1,value=name); I.cell(row=i,column=2,value=lo).fill=inp; I.cell(row=i,column=3,value=hi).fill=inp
    I.cell(row=i,column=4,value=unit); I.cell(row=i,column=5,value=lab); I.cell(row=i,column=6,value=note); R[k]=i
widths(I,[62,12,12,12,26,70]); I.freeze_panes="B2"
def ref(k,c): return f"Inputs!${'B' if c=='B' else 'C'}${R[k]}"

# ---------------- Derived ----------------
D=wb.create_sheet("Derived")
header(D,1,["Quantity","Low","High","Unit"])
der=[("Number of buildings",lambda c:f"={ref('sf_per_mw',c)}*{ref('it_mw',c)}/{ref('footprint_sf',c)}","count"),
("Exterior wall area, sf",lambda c:f"={c}2*{ref('wall_sf_bldg',c)}","sf"),
("Exterior wall area, m²",lambda c:f"={c}3*{SF_M2}","m²"),
("Fence face area, sf",lambda c:f"={ref('fence_km',c)}*1000*{ref('fence_h_m',c)}/{SF_M2}","sf"),
("SUPERWOOD board mass per sf",lambda c:f"={ref('board_mm',c)}/1000*{ref('sw_kg_m3',c)}*{SF_M2}","kg/sf"),
("SuperMill One output in tonnes",lambda c:f"={ref('sm1_sf',c)}*{c}6/1000","t/yr"),
("SuperMill Two output in tonnes",lambda c:f"={ref('sm2_sf',c)}*{c}6/1000","t/yr")]
for i,(n,f,u) in enumerate(der,start=2):
    D.cell(row=i,column=1,value=n); D.cell(row=i,column=4,value=u)
    for c,col in (("B",2),("C",3)): x=D.cell(row=i,column=col,value=f(c)); x.fill=calc; x.number_format="#,##0.00"
widths(D,[40,16,16,10])
def dref(row,c): return f"Derived!{c}{row}"

# ---------------- Foundations (footprint-based concrete with soil cases) ----------------
F=wb.create_sheet("Foundations")
header(F,1,["Element","Low  m³","High  m³","Basis"])
CFT="0.0283168"  # m³ per cubic foot
FR={}
fdef=[
("sf","Building floor area, sf",lambda c:f"={ref('sf_per_mw',c)}*{ref('it_mw',c)}","from Inputs"),
("perim","Total hall perimeter, ft",lambda c:f"={dref(2,c)}*4*SQRT({ref('footprint_sf',c)})","buildings x 4 sides"),
("cols","Columns",lambda c:f"={c}{{sf}}/({ref('bay_ft',c)}^2)","one per bay"),
("slab","Hall slab on grade",lambda c:f"={c}{{sf}}*{ref('slab_thk_in',c)}/12*{CFT}","6–8 in over the floor"),
("thick","Thickened slab zones",lambda c:f"={c}{{sf}}*{ref('thick_share',c)}*({ref('thick_thk_in',c)}-{ref('slab_thk_in',c)})/12*{CFT}","extra depth over 15–25% of the floor"),
("footings","Column footings",lambda c:f"={c}{{cols}}*{ref('footing_ft',c)}^2*{ref('footing_d_ft',c)}*{CFT}","spread footings, good soils"),
("gb","Perimeter grade beams",lambda c:f"={c}{{perim}}*{ref('gb_w_ft',c)}*{ref('gb_d_ft',c)}*{CFT}",""),
("gens","Generators",lambda c:f"={ref('it_mw',c)}/{ref('genset_mw',c)}*{ref('genset_redund',c)}","units incl. redundancy"),
("genpads","Generator pads",lambda c:f"={c}{{gens}}*{ref('genset_pad_cf',c)}*{CFT}",""),
("xfpads","Transformer and switchgear pads",lambda c:f"={c}{{gens}}*{ref('xfmr_per_gen',c)}*{ref('xfmr_pad_cf',c)}*{CFT}",""),
("mechpads","Mechanical equipment pads",lambda c:f"={c}{{sf}}*{ref('mech_pad_share',c)}*{ref('mech_pad_thk_in',c)}/12*{CFT}",""),
("caseA","Case 1 good soils — building and pads",lambda c:f"={c}{{slab}}+{c}{{thick}}+{c}{{footings}}+{c}{{gb}}+{c}{{genpads}}+{c}{{xfpads}}+{c}{{mechpads}}","sum"),
("lean","Case 2 — lean concrete / CLSM under footprint",lambda c:f"={c}{{sf}}*{ref('lean_share',c)}*{ref('lean_thk_in',c)}/12*{CFT}",""),
("uplift","Case 2 — larger footings, grade beams and pads",lambda c:f"=({c}{{footings}}+{c}{{gb}}+{c}{{genpads}}+{c}{{xfpads}}+{c}{{mechpads}})*{ref('found_uplift',c)}",""),
("caseB_allow","Case 2 moderate soils — allowance",lambda c:f"={c}{{lean}}+{c}{{uplift}}","added to case 1"),
("piles","Case 3 — piles or drilled piers under columns",lambda c:f"={c}{{cols}}*{ref('piles_per_col',c)}*PI()*({ref('pile_dia_ft',c)}/2)^2*{ref('pile_len_ft',c)}*{CFT}",""),
("caps","Case 3 — pile caps net of spread footings",lambda c:f"={c}{{cols}}*{ref('cap_cf',c)}*{CFT}-{c}{{footings}}",""),
("gbpiers","Case 3 — piers under grade beams",lambda c:f"={c}{{perim}}/{ref('gb_pier_spacing_ft',c)}*PI()*({ref('pile_dia_ft',c)}/2)^2*{ref('pile_len_ft',c)}*{CFT}",""),
("structslab","Case 3 — structural slab thickening",lambda c:f"={c}{{sf}}*{ref('struct_slab_add_in',c)}/12*{CFT}",""),
("caseC_allow","Case 3 poor soils — allowance (includes the case 2 measures)",lambda c:f"={c}{{caseB_allow}}+{c}{{piles}}+{c}{{caps}}+{c}{{gbpiers}}+{c}{{structslab}}","nested: ground replacement and larger elements plus piles, caps, piers and a structural slab"),
("allow","Selected soil allowance",lambda c:f"=CHOOSE({ref('soil_case',c)},0,{c}{{caseB_allow}},{c}{{caseC_allow}})","Inputs soil_case"),
("paving","Site paving, roads, yards, parking",lambda c:f"={c}{{sf}}*{ref('site_paving_ratio',c)}*{ref('site_paving_thk_in',c)}/12*{CFT}",""),
("yardmats","Substation / transformer yard mats",lambda c:f"={ref('yard_mat_acres',c)}*{ref('it_mw',c)}/1000*43560*{ref('yard_mat_thk_in',c)}/12*{CFT}",""),
("allin","All-in with site work, selected soil case",lambda c:f"={c}{{caseA}}+{c}{{allow}}+{c}{{paving}}+{c}{{yardmats}}","sum"),
("slab_basis_m3","Slab and paving, footprint basis (m³)",lambda c:f"={c}{{slab}}+{c}{{thick}}+{c}{{paving}}+IF({ref('soil_case',c)}=3,{c}{{structslab}},0)","feeds the main sheet when concrete_basis = 1"),
("found_basis_m3","Foundations, footings and pads, footprint basis (m³)",lambda c:f"={c}{{footings}}+{c}{{gb}}+{c}{{genpads}}+{c}{{xfpads}}+{c}{{mechpads}}+{c}{{yardmats}}+IF({ref('soil_case',c)}>=2,{c}{{caseB_allow}},0)+IF({ref('soil_case',c)}=3,{c}{{piles}}+{c}{{caps}}+{c}{{gbpiers}},0)","feeds the main sheet when concrete_basis = 1; soil cases nest"),
("slab_basis_t","Slab and paving, footprint basis (t)",lambda c:f"={c}{{slab_basis_m3}}*{ref('concrete_t_m3',c)}",""),
("found_basis_t","Foundations, footings and pads, footprint basis (t)",lambda c:f"={c}{{found_basis_m3}}*{ref('concrete_t_m3',c)}",""),
("permw_m3","Per-MW basis for comparison (m³)",lambda c:f"={ref('concrete_m3_mw',c)}*{ref('it_mw',c)}","published, secondary"),
]
for i,(k,n,f,basis) in enumerate(fdef,start=2): FR[k]=i
for i,(k,n,f,basis) in enumerate(fdef,start=2):
    F.cell(row=i,column=1,value=n); F.cell(row=i,column=4,value=basis)
    for c,col in (("B",2),("C",3)):
        formula=f(c)
        for kk,rr in FR.items(): formula=formula.replace("{"+kk+"}",str(rr))
        x=F.cell(row=i,column=col,value=formula); x.fill=calc; x.number_format="#,##0"
    if k in ("caseA","caseB_allow","caseC_allow","allin","slab_basis_t","found_basis_t"): F.cell(row=i,column=1).font=tot
F.cell(row=len(fdef)+3,column=1,value="Wall system: tilt-up versus insulated metal panel is the Inputs walls_precast switch (precast row on the main sheet, steel-share sheet reports both). Soil cases and the concrete basis switch are Inputs soil_case and concrete_basis. All element inputs are estimates [L]; see foundation-bottom-up.md.")
widths(F,[58,16,16,44])
def fref(k,c): return f"Foundations!{c}{FR[k]}"

# ---------------- 1 GW data center (main sheet) ----------------
SHEET="1 GW data center"
M=wb.create_sheet(SHEET,1)
HORS=("Immediate","Soon","Medium term","Long term"); HCOL={"Immediate":6,"Soon":7,"Medium term":8,"Long term":9}
header(M,1,["Component (building, then contents)","Low  tons per data center","High  tons per data center","Low  cumulative t","High  cumulative t",
            "Addressable — immediate","Addressable — soon","Addressable — medium term","Addressable — long term",
            "Low  replaced t","High  replaced t","Low  SUPERWOOD t","High  SUPERWOOD t","Gate / basis","High  not replaced t",
            "Low SW immediate","Low SW soon","Low SW medium","Low SW long","High SW immediate","High SW soon","High SW medium","High SW long",
            "Carbon factor (kg CO₂e/kg)","Low embodied carbon (t CO₂e)","High embodied carbon (t CO₂e)","EAF-basis factor (kg CO₂e/kg)","Material class","Steel share","Concrete share","Plastic share","Other share"])
import json as _json; SPLIT=_json.load(open("material_split.json"))["split"]
def split_for(name):
    n=name.replace("Structural steel — ","Steel — ").replace("Concrete: slab on grade, paving, yard","Concrete — slab on grade, paving").replace("Concrete: foundations, footings, piers, equipment pads","Concrete — foundations, footings, pads")
    for k,v in SPLIT.items():
        if n.split(" (")[0].startswith(k.split(" (")[0][:24]): return v
    return None
IT=lambda c: ref('it_mw',c)
# (name, mass formula, horizon, share formula, sw rule, note)
comp=[
("BUILDING — STRUCTURE AND ENVELOPE",None,None,None,None,None),
("Concrete: slab on grade, paving, yard",lambda c:f"=IF({ref('concrete_basis',c)}=1,{fref('slab_basis_t',c)},{ref('concrete_m3_mw',c)}*(1-{ref('foundation_share',c)})*{ref('concrete_t_m3',c)}*{IT(c)})-{{precast}}","Long term",lambda c:f"={ref('slab_lt',c)}","concrete","Technical potential per Alex (75%); no design or code pathway yet. arXiv 2509.21312 [M] for total concrete"),
("Concrete: foundations, footings, piers, equipment pads",lambda c:f"=IF({ref('concrete_basis',c)}=1,{fref('found_basis_t',c)},{ref('concrete_m3_mw',c)}*{ref('foundation_share',c)}*{ref('concrete_t_m3',c)}*{IT(c)})","Long term",lambda c:f"={ref('fdn_lt',c)}","concrete","Technical potential per Alex (90%); lightweight insulated SUPERWOOD foundations — geotechnical, durability and code pathway all open"),
("Precast / tilt-up perimeter walls (precast case)",lambda c:f"={ref('walls_precast',c)}*{dref(4,c)}*{ref('precast_thk_m',c)}*{ref('concrete_t_m3',c)}","Medium term",lambda c:"1","hybrid","Hybrid SUPERWOOD wall panels; NFPA 285 + E119 assemblies. Active when Inputs walls_precast = 1"),
("Rebar in all concrete",lambda c:f"=IF({ref('concrete_basis',c)}=1,{fref('slab_basis_m3',c)}+{fref('found_basis_m3',c)},{ref('concrete_m3_mw',c)}*{IT(c)})*{ref('rebar_kg_m3',c)}/1000","Long term",lambda c:f"={ref('foundation_share',c)}*{ref('fdn_lt',c)}+(1-{ref('foundation_share',c)})*{ref('slab_lt',c)}","steel","Rebar follows the concrete it sits in: foundation share × 90% + slab share × 75%"),
("Structural steel — primary frame (columns, girders)",lambda c:f"={ref('steel_t_mw',c)}*(1-{ref('secondary_share',c)})*{IT(c)}","Medium term",lambda c:"1","steel","ICC-ES via the mass-timber qualification pathway; FM acceptance; E119"),
("Structural steel — roof trusses, joists, roof deck, girts",lambda c:f"={ref('steel_t_mw',c)}*{ref('secondary_share',c)}*{IT(c)}","Medium term",lambda c:"1","steel","Design values and connection data; trusses carry no fire-resistance requirement"),
("Exterior skins — metal panel / IMP (metal-panel case)",lambda c:f"=(1-{ref('walls_precast',c)})*{dref(4,c)}*{ref('panel_kg_m2',c)}/1000","Immediate",lambda c:"1","skins","Shipping now; area-for-area rain screen"),
("Louvers and yard / mechanical screens",lambda c:f"={ref('yard_screen_sf',c)}*{SF_M2}*{ref('louver_kg_m2',c)}/1000","Immediate",lambda c:"1","louvers","Shipping now"),
("Security and staff-area fencing",lambda c:f"={ref('fence_km',c)}*1000*{ref('fence_kg_m',c)}/1000","Immediate",lambda c:"1","fence","Shipping now — the Microsoft near-term list"),
("CONTENTS — FIT-OUT AND EQUIPMENT",None,None,None,None,None),
("Platforms, walkways, mezzanines, railings, tray supports",lambda c:f"={ref('platform_t_mw',c)}*{IT(c)}","Soon",lambda c:"1","steel","Published design values; IBC 1607 for railings"),
("Acoustic barriers, enclosures, HVAC separations",lambda c:f"={ref('acoustic_t_mw',c)}*{IT(c)}","Soon",lambda c:"1","steel","STC / OITC lab and field data"),
("Racking and equipment supports",lambda c:f"={ref('racking_t_mw',c)}*{IT(c)}","Soon",lambda c:"1","steel","A few months of design and load-data development (Alex 2026-09-04)"),
("Other tray, containment, doors, misc. metals",lambda c:f"={ref('other_t_mw',c)}*{IT(c)}","Soon",lambda c:f"={ref('share_other',c)}","steel","Partial; UL 10C for rated doors"),
("Ducting, plenums and air-distribution sheet metal",lambda c:f"={ref('duct_t_mw',c)}*{IT(c)}","Medium term",lambda c:"1","steel","NFPA 90A / UL 181 noncombustibility expectations for in-airstream components — the hardest gate in this row set"),
("Interior finishes, backplanes, trim (admin / office)",lambda c:f"={ref('interior_t_mw',c)}*{IT(c)}","Immediate",lambda c:"1","interior","E84 Class A finish; backplanes UL 94 yellow card (not yet started)"),
("Electrical equipment and conductors",lambda c:f"={ref('elec_t_mw',c)}*{IT(c)}",None,lambda c:"0","zero","Gensets, transformers, switchgear, batteries, copper"),
("Mechanical equipment, piping, loop water",lambda c:f"={ref('mech_t_mw',c)}*{IT(c)}",{"Medium term":lambda c:f"={ref('mech_shell_share',c)}","Long term":lambda c:f"={ref('mech_blade_share',c)}"},None,"steel","Shells, casings and fan-wall frames medium term (Alex 2026-09-06); fan blades long term, needs investigation; compressors, pumps, tubes, piping and water stay"),
("IT — servers and racks",lambda c:f"={ref('it_t_mw',c)}*{IT(c)}",{"Soon":lambda c:f"={ref('it_rack_share',c)}","Long term":lambda c:f"={ref('it_enclosure_share',c)}"},None,"steel","Rack cabinets soon (18% of IT mass), server enclosures long term (30%), per analyses/it-rack-material-stack.md; electronics never. Rack masses [H]; aggregate [L]"),
]
r=2; data_rows=[]; precast_row=None
for name,f,hor,share,rule,note in comp:
    if f is None:
        M.cell(row=r,column=1,value=name).font=tot; M.cell(row=r,column=1).fill=sect; r+=1; continue
    if name.startswith("Precast"): precast_row=r
    data_rows.append(r); r+=1
r=2
for name,f,hor,share,rule,note in comp:
    if f is None: r+=1; continue
    M.cell(row=r,column=1,value=name); M.cell(row=r,column=14,value=note)
    for c,col in (("B",2),("C",3)):
        M.cell(row=r,column=col,value=f(c).replace("{precast}",f"{c}{precast_row}")).fill=calc
        prev=[x for x in data_rows if x<r]
        M.cell(row=r,column=col+2,value=f"={c}{r}" if not prev else f"={CL(col+2)}{prev[-1]}+{c}{r}").fill=calc
    # four addressable-share columns (editable); the row's horizon gets the share, others 0
    for h in HORS:
        _s=(hor[h]("B") if h in hor else "0") if isinstance(hor,dict) else (share("B") if h==hor else "0"); _s=float(_s) if not _s.startswith("=") else _s
        x=M.cell(row=r,column=HCOL[h],value=_s); x.fill=inp; x.number_format="0%"
    tot_share=f"SUM($F{r}:$I{r})"
    for c,(col,scol) in (("B",(10,12)),("C",(11,13))):
        M.cell(row=r,column=col,value=f"={c}{r}*{tot_share}").fill=calc
        repl=f"{CL(col)}{r}"
        if rule=="zero": sw="=0"
        elif rule=="steel": sw=f"={repl}*{ref('sub_factor',c)}"
        elif rule=="concrete": sw=f"={repl}*{ref('conc_sub_factor',c)}"
        elif rule=="skins": sw=f"={tot_share}*(1-{ref('walls_precast',c)})*{dref(3,c)}*{dref(6,c)}/1000"
        elif rule=="louvers": sw=f"={tot_share}*{ref('yard_screen_sf',c)}*{dref(6,c)}/1000"
        elif rule=="fence": sw=f"={tot_share}*{dref(5,c)}*{dref(6,c)}/1000"
        elif rule=="hybrid": sw=f"={tot_share}*{ref('walls_precast',c)}*{dref(4,c)}*{ref('hybrid_kg_m2',c)}/1000"
        elif rule=="interior": sw=f"={tot_share}*({ref('interior_sf',c)}+{ref('backplane_sf',c)})*{dref(6,c)}/1000"
        M.cell(row=r,column=scol,value=sw).fill=calc
        # helper: SUPERWOOD split by horizon in proportion to the shares
        base=16 if c=="B" else 20
        for j,h in enumerate(HORS):
            M.cell(row=r,column=base+j,value=f"=IF({tot_share}=0,0,{CL(scol)}{r}*{CL(HCOL[h])}{r}/{tot_share})").fill=calc
    M.cell(row=r,column=15,value=f"=C{r}-K{r}").fill=calc
    # embodied carbon per row: steel, concrete and plastic content at their factors; "other" (copper, aluminum, water, gypsum, wood, electronics) carries no factor (Alex 2026-09-05)
    kind = "concrete" if (name.startswith("Concrete") or name.startswith("Precast")) else "equipment" if (name.startswith("Electrical") or name.startswith("Mechanical") or name.startswith("IT")) else "mixed" if rule=="interior" else "steel"
    M.cell(row=r,column=28,value=kind)
    sp=split_for(name) or ([0,1,0,0] if kind=="concrete" else [1,0,0,0])
    for j,v in enumerate(sp): M.cell(row=r,column=29+j,value=v).fill=inp
    M.cell(row=r,column=24,value=f"=AC{r}*{ref('co2_steel','B')}+AD{r}*{ref('co2_conc','B')}+AE{r}*{ref('co2_plastic','B')}").fill=calc
    M.cell(row=r,column=27,value=f"=AC{r}*{ref('co2_eaf','C')}+AD{r}*{ref('co2_conc','B')}+AE{r}*{ref('co2_plastic','B')}").fill=calc
    M.cell(row=r,column=25,value=f"=B{r}*X{r}").fill=calc; M.cell(row=r,column=26,value=f"=C{r}*X{r}").fill=calc
    r+=1
first,last=data_rows[0],data_rows[-1]
r+=1
def srow(label,fB,fC,fmt="#,##0",bold=False):
    global r
    M.cell(row=r,column=1,value=label).font=Font(bold=bold)
    x=M.cell(row=r,column=2,value=fB); x.number_format=fmt; x.font=Font(bold=bold)
    y=M.cell(row=r,column=3,value=fC); y.number_format=fmt; y.font=Font(bold=bold)
    r+=1; return r-1
M.cell(row=r,column=1,value="SUMMARY — per data center").font=tot; M.cell(row=r,column=1).fill=sect; r+=1
rt=srow("Total mass, building and contents (tons)",f"=SUM(B{first}:B{last})",f"=SUM(C{first}:C{last})",bold=True)
rx=srow("Total excluding concrete (tons)",f"=B{rt}-B{first}-B{first+1}-B{precast_row}",f"=C{rt}-C{first}-C{first+1}-C{precast_row}",bold=True)
r+=1
M.cell(row=r,column=1,value="Replaced by SUPERWOOD, by horizon").font=tot; r+=1
hor_rows={}
for j,hor in enumerate(HORS):
    sc=CL(HCOL[hor])
    rr=srow(f"{hor} — incumbent mass replaced (tons)",f"=SUMPRODUCT(B{first}:B{last},{sc}{first}:{sc}{last})",f"=SUMPRODUCT(C{first}:C{last},{sc}{first}:{sc}{last})")
    rs=srow(f"{hor} — SUPERWOOD required (tons)",f"=SUM({CL(16+j)}{first}:{CL(16+j)}{last})",f"=SUM({CL(20+j)}{first}:{CL(20+j)}{last})")
    if hor=="Immediate":
        srow("Immediate — SUPERWOOD required (sf)",f"=B{rs}*1000/Derived!B6",f"=C{rs}*1000/Derived!C6")
        ry=srow("Immediate — years of SuperMill One output",f"=B{rs}/Derived!B7",f"=C{rs}/Derived!C7",fmt="0.0")
    else:
        ry=srow(f"{hor} — years of SuperMill Two output",f"=B{rs}/Derived!B8",f"=C{rs}/Derived!C8",fmt="0.0")
    hor_rows[hor]=(rr,rs,ry); r+=1
rc=srow("Cumulative incumbent mass replaced (tons)",f"=SUM(J{first}:J{last})",f"=SUM(K{first}:K{last})",bold=True)
rsw=srow("Cumulative SUPERWOOD required (tons)",f"=SUM(L{first}:L{last})",f"=SUM(M{first}:M{last})",bold=True)
rsy=srow("Cumulative years of SuperMill Two output",f"=B{rsw}/Derived!B8",f"=C{rsw}/Derived!C8",fmt="0.0")
rsh=srow("Share of total mass replaced",f"=B{rc}/B{rt}",f"=C{rc}/C{rt}",fmt="0.0%")
rshx=srow("Share of ex-concrete mass replaced through the medium term (excludes foundations)",f"=(B{rc}-B{hor_rows['Long term'][0]})/B{rx}",f"=(C{rc}-C{hor_rows['Long term'][0]})/C{rx}",fmt="0.0%")
rnr=srow("Not replaced by SUPERWOOD (tons)",f"=B{rt}-B{rc}",f"=C{rt}-C{rc}",bold=True)
rnrs=srow("Not replaced, share of total",f"=B{rnr}/B{rt}",f"=C{rnr}/C{rt}",fmt="0.0%")
r+=1
M.cell(row=r,column=1,value="EMBODIED CARBON — building materials (equipment not estimated)").font=tot; M.cell(row=r,column=1).fill=sect; r+=1
rcb=srow("Embodied carbon, building materials (t CO₂e)",f"=SUM(Y{first}:Y{last})",f"=SUM(Z{first}:Z{last})",bold=True)
rcs=srow("  of which steel (t CO₂e)",f'=SUMIF(AB{first}:AB{last},"steel",Y{first}:Y{last})',f'=SUMIF(AB{first}:AB{last},"steel",Z{first}:Z{last})')
rcc=srow("  of which concrete (t CO₂e)",f'=SUMIF(AB{first}:AB{last},"concrete",Y{first}:Y{last})',f'=SUMIF(AB{first}:AB{last},"concrete",Z{first}:Z{last})')
rcss=srow("Steel share of building-materials carbon",f"=B{rcs}/B{rcb}",f"=C{rcs}/C{rcb}",fmt="0.0%")
rccs=srow("Concrete share of building-materials carbon",f"=B{rcc}/B{rcb}",f"=C{rcc}/C{rcb}",fmt="0.0%")
rebar_row=[x for x in data_rows if M.cell(row=x,column=1).value.startswith("Rebar")][0]
rcas=srow("Steel above the slab (steel excl. rebar), share of building-materials carbon",f"=(B{rcs}-Y{rebar_row})/B{rcb}",f"=(C{rcs}-Z{rebar_row})/C{rcb}",fmt="0.0%")
rceaf=srow("Steel share if steel is recycled (EAF factor, high end)",f'=SUMPRODUCT(B{first}:B{last},AA{first}:AA{last},--(AB{first}:AB{last}="steel"))/SUMPRODUCT(B{first}:B{last},AA{first}:AA{last})',f'=SUMPRODUCT(C{first}:C{last},AA{first}:AA{last},--(AB{first}:AB{last}="steel"))/SUMPRODUCT(C{first}:C{last},AA{first}:AA{last})',fmt="0.0%")
for rr in range(2,last+1):
    for col in (2,3,4,5,10,11,12,13,15)+tuple(range(16,24)):
        if M.cell(row=rr,column=col).number_format=="General": M.cell(row=rr,column=col).number_format="#,##0"
widths(M,[58,15,15,16,16,12,12,12,12,14,14,14,14,64,16]+[11]*8+[14,18,18,14,12,10,10,10,10]); M.freeze_panes="B2"
for rr in range(first,last+1):
    for col in (25,26): M.cell(row=rr,column=col).number_format="#,##0"
    M.cell(row=rr,column=24).number_format="0.00"; M.cell(row=rr,column=27).number_format="0.00"
for col in range(16,24): M.column_dimensions[CL(col)].outlineLevel=1; M.column_dimensions[CL(col)].hidden=True
M.cell(row=1,column=16).value="Low SW immediate (helper — grouped, unhide to see the per-horizon split)"
ch=BarChart(); ch.type="bar"; ch.grouping="stacked"; ch.overlap=100; ch.title="High case, tons per data center — replaced by SUPERWOOD vs not (slab concrete row excluded)"
rb=first
ch.add_data(Reference(M,min_col=11,min_row=rb,max_row=last),titles_from_data=False)
ch.add_data(Reference(M,min_col=15,min_row=rb,max_row=last),titles_from_data=False)
ch.set_categories(Reference(M,min_col=1,min_row=rb,max_row=last))
from openpyxl.chart.series import SeriesLabel
ch.series[0].title=SeriesLabel(v="Replaced by SUPERWOOD (high)"); ch.series[1].title=SeriesLabel(v="Not replaced (high)")
ch.series[0].graphicalProperties.solidFill="E2B877"; ch.series[1].graphicalProperties.solidFill="9D8D76"
ch.height=12; ch.width=26; ch.y_axis.title="tons per data center"
M.add_chart(ch,f"A{r+2}")

# ---------------- Carbon ----------------
C=wb.create_sheet("Carbon")
header(C,1,["Horizon","Low incumbent emissions avoided (t CO₂e)","High","Low SUPERWOOD manufacturing (t CO₂e)","High","Low net reduction","High","Low net vs EAF steel","High","Low biogenic stored (separate)","High"])
for i,hor in enumerate(("Immediate","Soon","Medium term","Long term"),start=2):
    rr,rs,_=hor_rows[hor]; C.cell(row=i,column=1,value=hor+(" (concrete, rebar, server enclosures)" if hor=="Long term" else " (steel)"))
    for c,col in (("B",2),("C",3)):
        sw=f"'1 GW data center'!{c}{rs}"; sc=CL(HCOL[hor]); S=f"'1 GW data center'!"
        C.cell(row=i,column=col,value=f"=SUMPRODUCT({S}{c}{first}:{c}{last},{S}{sc}{first}:{sc}{last},{S}X{first}:X{last})")
        C.cell(row=i,column=col+2,value=f"={sw}*{ref('co2_sw',c)}")
        C.cell(row=i,column=col+4,value=f"={CL(col)}{i}-{CL(col+2)}{i}")
        C.cell(row=i,column=col+6,value=f"=SUMPRODUCT({S}{c}{first}:{c}{last},{S}{sc}{first}:{sc}{last},{S}AA{first}:AA{last})-{CL(col+2)}{i}")
        C.cell(row=i,column=col+8,value=f"={sw}*{ref('co2_bio',c)}")
    for col in range(2,12): C.cell(row=i,column=col).number_format="#,##0"
C.cell(row=7,column=1,value="Pre-LCA projections (LCA under way with Prof. Ming Hu, University of Notre Dame). Each row is valued at its own factor (column X on the main sheet): concrete rows at co2_conc, steel rows including rebar at co2_steel, interior finishes at co2_mixed; equipment rows (including the long-term server enclosures) carry no factor, so the long-term row counts concrete and rebar only. The EAF column revalues steel rows at the recycled-steel factor (high end). Equipment rows carry no factor. Biogenic storage reported separately (EN 15804 module C).")
widths(C,[30,18,12,18,12,16,12,16,12,18,12])

# ---------------- Steel share of above-ground mass ----------------
SS=wb.create_sheet("Steel share")
header(SS,1,["Above-ground component (contents included)","Low tons","High tons","Steel fraction — low","Steel fraction — high","Low steel tons","High steel tons","Basis for the fraction"])
fr={"Structural steel — primary":(1,1,"steel by definition"),"Structural steel — roof":(1,1,"steel by definition"),"Exterior skins":(0.85,0.95,"steel-faced IMP; some aluminum"),
"Louvers":(0.4,0.7,"aluminum common"),"Security":(0.9,1,"chain-link, palisade, posts"),"Platforms":(0.95,1,"structural and misc. steel"),"Acoustic":(0.7,0.9,"steel panels with absorptive fill"),
"Racking":(0.95,1,"steel"),"Other tray":(0.85,0.95,"mostly steel, some aluminum"),"Ducting":(0.95,1,"galvanized steel"),"Interior finishes":(0.2,0.4,"gypsum, wood, steel studs"),
"Electrical":(0.5,0.65,"enclosures, cores, gensets, switchgear; rest copper, oil, batteries"),"Mechanical":(0.45,0.65,"chillers, piping steel; water, copper, refrigerant not"),"IT":(0.4,0.6,"steel chassis and racks; aluminum, PCBs, copper"),
"Precast":(0.03,0.05,"rebar in precast/tilt-up panels; row is zero unless Inputs walls_precast = 1")}
rr=2; ss_rows=[]
for x in data_rows:
    nm=M.cell(row=x,column=1).value
    if nm.startswith("Concrete") or nm.startswith("Rebar"): continue
    key=[k for k in fr if nm.startswith(k)]
    if not key: continue
    lo_f,hi_f,basis=fr[key[0]]
    SS.cell(row=rr,column=1,value=nm); SS.cell(row=rr,column=2,value=f"='{SHEET}'!B{x}").fill=calc; SS.cell(row=rr,column=3,value=f"='{SHEET}'!C{x}").fill=calc
    a=SS.cell(row=rr,column=4,value=lo_f); a.fill=inp; a.number_format="0%"; bb=SS.cell(row=rr,column=5,value=hi_f); bb.fill=inp; bb.number_format="0%"
    SS.cell(row=rr,column=6,value=f"=B{rr}*D{rr}").fill=calc; SS.cell(row=rr,column=7,value=f"=C{rr}*E{rr}").fill=calc; SS.cell(row=rr,column=8,value=basis)
    for col in (2,3,6,7): SS.cell(row=rr,column=col).number_format="#,##0"
    ss_rows.append(rr); rr+=1
f1,f2=ss_rows[0],ss_rows[-1]; rr+=1
SS.cell(row=rr,column=1,value="Above-ground mass, contents included (tons)").font=Font(bold=True); SS.cell(row=rr,column=2,value=f"=SUM(B{f1}:B{f2})"); SS.cell(row=rr,column=3,value=f"=SUM(C{f1}:C{f2})"); ss_tot=rr; rr+=1
SS.cell(row=rr,column=1,value="Steel in it (tons)").font=Font(bold=True); SS.cell(row=rr,column=2,value=f"=SUM(F{f1}:F{f2})"); SS.cell(row=rr,column=3,value=f"=SUM(G{f1}:G{f2})"); ss_steel=rr; rr+=1
SS.cell(row=rr,column=1,value="Steel share of above-ground mass").font=Font(bold=True); SS.cell(row=rr,column=2,value=f"=B{ss_steel}/B{ss_tot}"); SS.cell(row=rr,column=3,value=f"=C{ss_steel}/C{ss_tot}"); ss_share=rr
for col in (2,3): SS.cell(row=rr,column=col).number_format="0%"
rr+=1
SS.cell(row=rr,column=1,value="Steel share including slab rebar (at-grade steel)").font=Font(bold=True)
SS.cell(row=rr,column=2,value=f"=(B{ss_steel}+'{SHEET}'!B{rebar_row})/(B{ss_tot}+'{SHEET}'!B{rebar_row})"); SS.cell(row=rr,column=3,value=f"=(C{ss_steel}+'{SHEET}'!C{rebar_row})/(C{ss_tot}+'{SHEET}'!C{rebar_row})"); ss_share_rebar=rr
for col in (2,3): SS.cell(row=rr,column=col).number_format="0%"
rr+=1
SS.cell(row=rr,column=1,value="Structure and envelope only (first five rows)").font=Font(bold=True)
SS.cell(row=rr,column=2,value=f"=SUM(F{f1}:F{f1+5})/SUM(B{f1}:B{f1+5})"); SS.cell(row=rr,column=3,value=f"=SUM(G{f1}:G{f1+5})/SUM(C{f1}:C{f1+5})"); ss_se=rr
for col in (2,3): SS.cell(row=rr,column=col).number_format="0%"
rr+=2
SS.cell(row=rr,column=1,value="WALL SYSTEM CASES (independent of the Inputs toggle)").font=tot; SS.cell(row=rr,column=1).fill=sect; rr+=1
pre_row=[x for x in ss_rows if SS.cell(row=x,column=1).value.startswith("Precast")][0]; skin_row=[x for x in ss_rows if SS.cell(row=x,column=1).value.startswith("Exterior")][0]
# forced precast mass and metal-skin mass, regardless of the flag
SS.cell(row=rr,column=1,value="Tilt-up / precast wall mass if chosen (tons)"); SS.cell(row=rr,column=2,value=f"=Derived!B4*{ref('precast_thk_m','B')}*{ref('concrete_t_m3','B')}").fill=calc; SS.cell(row=rr,column=3,value=f"=Derived!C4*{ref('precast_thk_m','C')}*{ref('concrete_t_m3','C')}").fill=calc; ss_pre=rr; rr+=1
SS.cell(row=rr,column=1,value="Metal-panel skin mass if chosen (tons)"); SS.cell(row=rr,column=2,value=f"=Derived!B4*{ref('panel_kg_m2','B')}/1000").fill=calc; SS.cell(row=rr,column=3,value=f"=Derived!C4*{ref('panel_kg_m2','C')}/1000").fill=calc; ss_skin=rr; rr+=1
# metal-panel case: total and steel with the skin row forced in and precast forced out
SS.cell(row=rr,column=1,value="Steel share — metal-panel walls").font=Font(bold=True)
for c,col in (("B",2),("C",3)):
    tot_=f"({c}{ss_tot}-{c}{pre_row}-{c}{skin_row}+{c}{ss_skin})"; st_=f"({c}{ss_steel}-F{pre_row}-F{skin_row}+{c}{ss_skin}*{ 'D' if c=='B' else 'E'}{skin_row})".replace("F{pre_row}",f"{'F' if c=='B' else 'G'}{pre_row}").replace("F{skin_row}",f"{'F' if c=='B' else 'G'}{skin_row}")
    SS.cell(row=rr,column=col,value=f"={st_}/{tot_}").number_format="0%"
ss_mp=rr; rr+=1
SS.cell(row=rr,column=1,value="Steel share — tilt-up / precast concrete walls").font=Font(bold=True)
for c,col in (("B",2),("C",3)):
    tot_=f"({c}{ss_tot}-{c}{pre_row}-{c}{skin_row}+{c}{ss_pre})"; st_=f"({c}{ss_steel}-{'F' if c=='B' else 'G'}{pre_row}-{'F' if c=='B' else 'G'}{skin_row}+{c}{ss_pre}*{'D' if c=='B' else 'E'}{pre_row})"
    SS.cell(row=rr,column=col,value=f"={st_}/{tot_}").number_format="0%"
ss_tu=rr; rr+=2
SS.cell(row=rr,column=1,value="Excludes slab on grade, paving and foundations. The wall system matters: the main block follows Inputs walls_precast (0 = metal panel, 1 = tilt-up/precast); the two case rows above force each wall system regardless of the toggle. Steel fractions per component are estimates [conf: L] (yellow cells). Printed band on the deck: 50–80% (Alex, 2026-09-04), covering both wall systems.")
for col in (2,3,6,7): SS.cell(row=ss_tot,column=col).number_format="#,##0"; SS.cell(row=ss_steel,column=col).number_format="#,##0"
widths(SS,[58,14,14,16,16,14,14,64]); SS.freeze_panes="B2"

# ---------------- README ----------------
Rd=wb.create_sheet("README",0)
txt=["SUPERWOOD × Data Centers — material mass, embodied carbon and replacement model (2026-09-04, v3)",
"","Reference unit: 1 GW IT-load data center. Change any yellow cell on Inputs (and the four addressable-share columns on 1 GW data center — immediate, soon, medium term, long term; a row may split across them); everything recalculates.",
"Low column = every low input; High column = every high input. These are scenario bounds, not a distribution — vary single inputs for sensitivities.",
"","Read the sheet 1 GW data center top to bottom: building structure and envelope first, then contents, with a running cumulative total. The right-hand columns show, for each component, the addressable share in each of the four horizons, the mass replaced, and the SUPERWOOD required (per-horizon SUPERWOOD split sits in grouped helper columns P–W). The summary block and chart follow.",
"Horizons: Immediate (shipping now, no assembly rating) · Soon (certification-gated non-structural: platforms, railings, barriers, racking, doors) · Medium term (structural steel, roof trusses and roofs, ducting, enclosures) · Long term (foundations). Switch Inputs walls_precast to 1 to model precast perimeter walls; foundation_share and conc_sub_factor drive the long-term row.",
"","Provenance: concrete and structural-steel intensities are published, secondary (arXiv 2509.21312 citing Hasan 2022 / Sharma 2023, conf M); GB200 rack mass is NVIDIA-published (conf H). Everything else is estimated [conf: L] — replace with a real material takeoff from HITT, Turner or Fast + Epp before external use.",
"Embodied carbon: columns X–AB on the main sheet value each row at its own factor (concrete, steel, mixed; equipment not estimated), with totals and steel/concrete shares in the summary; the Carbon sheet rolls avoided emissions up by horizon at those factors and at the recycled-steel (EAF) factor. Steel share: the Steel share sheet estimates what fraction of above-ground mass, contents included, is steel (per-component steel fractions are yellow inputs).",
"Companion narrative: materials-mass-and-replacement.md in the same folder."]
for i,s in enumerate(txt,1): Rd.cell(row=i,column=1,value=s)
Rd.column_dimensions["A"].width=150
wb.save("materials-mass-and-replacement.xlsx"); print("saved; data rows",first,last,"summary ends",r)
import json; json.dump({"FR_names":{F.cell(row=r,column=1).value:r for r in FR.values()},"R":R,"sheet":SHEET,"first":first,"last":last,"rt":rt,"rx":rx,"hor":hor_rows,"rc":rc,"rsw":rsw,"rsy":rsy,"rsh":rsh,"rshx":rshx,"rnr":rnr,"rnrs":rnrs,"data_rows":data_rows,"precast":precast_row,"rcb":rcb,"rcs":rcs,"rcc":rcc,"rcss":rcss,"rccs":rccs,"rcas":rcas,"rceaf":rceaf,"ss_rows":ss_rows,"ss_tot":ss_tot,"ss_steel":ss_steel,"ss_share":ss_share,"ss_share_rebar":ss_share_rebar,"ss_se":ss_se,"ss_mp":ss_mp,"ss_tu":ss_tu,"ss_pre":ss_pre},open("model_rows.json","w"))
